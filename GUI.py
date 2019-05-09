import tkinter as tk
from tkinter import ttk
from openpyxl import Workbook
from openpyxl.styles import Border, Alignment, Color, Font, Side#, PatternFill, GradientFill
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.chart import BarChart, Reference #, Series
from tkinter import filedialog
from re import search as re_search
import sys
from os import system as os_system
from functools import partial

class Klassenliste:
    def __init__(self):

        self.gui_columns_width = [15,25,15]
        self.font_normal = ('Sans','12')
        self.font_bold = ('Sans','12','bold')
        self.font_small = ('Sans',str(int(self.font_normal[1])-2))

        self.root = tk.Tk()
        self.root.title('Excelersteller')
        self.container = tk.Frame(master=self.root, width=self.gui_columns_width[0])
        self.container.grid(column=0, row=0, sticky="WE", padx=15, pady=15)

        self.label_klassenliste = ttk.Label(master=self.container, text="Klassenliste:", font=self.font_normal)
        self.label_klassenliste.grid(column=0, row=0, sticky="W")

        self.klassenliste_txt = tk.StringVar()
        self.klassenliste_entered = tk.Entry(self.container, width=self.gui_columns_width[1], textvariable=self.klassenliste_txt, font=self.font_normal)
        self.klassenliste_entered.grid(column=1, row=0, sticky="W")
        self.Klassenliste_ok = False
        #self.klassenliste_entered.focus()

        self.bbutton= tk.Button(self.container, text="Browse", width=self.gui_columns_width[2], command=self.browse_txt, font=self.font_small)
        self.bbutton.grid(column=2, row=0, sticky="W", padx=10)

        self.label_benennung = ttk.Label(master=self.container, text="Bezeichnung:", width=self.gui_columns_width[0], font=self.font_normal)
        self.label_benennung.grid(column=0, row=1, sticky="W", pady=5) 
        self.benennung = tk.StringVar()
        self.benennung_entered = ttk.Entry(self.container, width=self.gui_columns_width[1], textvariable=self.benennung, font=self.font_normal)
        self.benennung_entered.grid(column=1, row=1, sticky="W", pady=5)

        self.label_aufgaben = ttk.Label(master=self.container, text="Aufgaben:", width=self.gui_columns_width[0], font=self.font_normal)
        self.label_aufgaben.grid(column=0, row=2, sticky="nw", pady=5) 
        self.task_container = tk.Frame(master=self.container)
        self.task_container.grid(column=1, row=2, sticky="NW")

        self.addTask = tk.Button(self.container, text="+", command=self.addBox, font=self.font_bold)
        self.addTask.grid(column=2, row=2, sticky="NW", padx=10, pady=5)

        self.btd_start = tk.Button(self.container, text='Excel erstellen', command=self.createExcel, font=self.font_normal)
        self.btd_start.grid(column=1, row=3, sticky="E", columnspan=2)

        self.logwindow = tk.Label(master=self.container, text='')
        self.logwindow.grid(column=0, row=4, sticky="W", columnspan=3)

        # Globale Strings definieren
        self.FIRSTNAME = 'Vorname'
        self.LASTNAME = 'Nachname'

        self.root.mainloop()


    def browse_txt(self):
        self.file_path = filedialog.askopenfilenames(title='Klassenliste auswählen... ', filetypes = (("txt files","*.txt"),("all files","*.*")))
        print(self.file_path)
        if self.file_path == '':
            self.klassenliste_txt.set('Keine Datei ausgewählt.')
            self.klassenliste_entered.config(foreground='red')
            self.Klassenliste_ok = False
        elif len(self.file_path) > 1:
            self.klassenliste_txt.set('Zu viele Dateien ausgewählt.')
            self.klassenliste_entered.config(foreground='red')
            self.Klassenliste_ok = False
        else:
            self.txtFile = re_search(r'/(\w*\d*\w*.txt)',self.file_path[0]).group(1)
            self.klassenliste_txt.set(self.txtFile)
            self.klassenliste_entered.config(foreground='green')
            self.Klassenliste_ok = True

    def addBox(self):

        self.be_width = 12
        self.entry_width = 10

        self.frame = tk.LabelFrame(self.task_container, text='')
        self.frame.grid(sticky="NW")

        self.tkvar = tk.StringVar(self.task_container)
        # Dictionary with options
        self.choices = { 'Aufgabe','Textproduktion'}
        self.tkvar.set('Aufgabe') # set the default options

        self.be_label = ttk.Label(self.frame, text='BE', font=self.font_normal, width=self.be_width)
        self.be_label.grid(column=0, row=1, sticky="W")
        self.be_var = tk.StringVar()
        self.be = tk.Entry(self.frame, textvariable=self.be_var, font=self.font_normal, width=self.entry_width)
        self.be.grid(column=1, row=1, sticky="W") 

        self.be2_label = ttk.Label(self.frame, text='', font=self.font_normal, width=self.be_width)
        self.be2_label.grid(column=0, row=2, sticky="W")
        self.be2_var = tk.StringVar()
        self.be2 = tk.Entry(self.frame, textvariable=self.be2_var, font=self.font_normal, width=self.entry_width)
        self.be2.grid(column=1, row=2, sticky="W") 
        self.be2.config(state='disabled')

        self.be3_label = ttk.Label(self.frame, text='', font=self.font_normal, width=self.be_width)
        self.be3_label.grid(column=0, row=3, sticky="w")
        self.be3_var = tk.StringVar()
        self.be3 = tk.Entry(self.frame, textvariable=self.be3_var, font=self.font_normal, width=self.entry_width)
        self.be3.grid(column=1, row=3, sticky="W")
        self.be3.config(state='disabled')


        self.btd = tk.Button(self.frame, text='X', fg='red', command=lambda x=len(all_entries): self.removeItem(x))
        self.btd.grid(row=0, column=2, sticky="N") 
        self.btd.config(font=self.font_bold)

        #self.choice = tk.OptionMenu(self.frame, self.tkvar, command=lambda x=len(all_entries): self.modifyTaskInputs(x), *self.choices)
        self.choice = tk.OptionMenu(self.frame, self.tkvar, *self.choices, command=partial(self.modifyTaskInputs,len(all_entries)))
        self.choice.grid(row=0, column=0, columnspan=2, sticky="W")
        self.choice.config(width=self.gui_columns_width[1]-4, font=self.font_normal)
        all_entries.append( (self.frame,self.tkvar,((self.be_label, self.be_var, self.be),(self.be2_label, self.be2_var, self.be2),(self.be3_label, self.be3_var, self.be3)),self.btd) )  

    def modifyTaskInputs(self,item,value):

        if value=='Textproduktion':
            for i in range(3):
                all_entries[item][2][i][2].config(state='normal')
            all_entries[item][2][0][0].config(text='Inhalt')
            all_entries[item][2][1][0].config(text='Sprache')
            all_entries[item][2][2][0].config(text='Gewichtung')
        elif value == "Aufgabe":
            all_entries[item][2][0][2].config(state='normal')
            for i in range(1,3):
                all_entries[item][2][i][2].config(state='disabled')
            all_entries[item][2][0][0].config(text='BE')
            all_entries[item][2][1][0].config(text='')
            all_entries[item][2][2][0].config(text='')
               

    def removeItem(self,item):
        # Frame entfernen
        all_entries[item][0].grid_forget()
        all_entries[item][0].destroy()
        all_entries.pop(item)

        # Die Befehle der verbleibenden Buttons an die Position anpassen
        for i in range(len(all_entries)):
            all_entries[i][-1].config(command=lambda x=len(all_entries)-1: self.removeItem(x))

    def createExcel(self):

        if self.klassenliste_txt.get() == '':
            self.logwindow.config(text='Keine Klassenliste ausgewählt', foreground='red')
        elif len(all_entries) == 0:
            self.logwindow.config(text='Keine Aufgaben angelegt', foreground="red")

        # wenn alles in ordnung ist
        elif self.klassenliste_txt.get() != '' and self.Klassenliste_ok:

            self.tasks = []
            for i in range(len(all_entries)):
                if all_entries[i][1].get() == 'Aufgabe':
                    self.tasks.append(Aufgabe(int(all_entries[i][2][0][1].get()),1))
                elif all_entries[i][1].get() == 'Textproduktion':
                    self.inhalt = int(all_entries[i][2][0][1].get())
                    self.sprache = int(all_entries[i][2][1][1].get())
                    self.gewichtung = int(all_entries[i][2][2][1].get())
                    self.tasks.append(Textproduktion(self.inhalt, self.sprache, self.gewichtung))

            self.logwindow.config(text='Excel wird erstellt...', foreground='blue')
            SUCCESS = self.runExcel()
            if SUCCESS == 0:
                self.logwindow.config(text='Excel wurde erfolgreich erstellt', foreground='green')
            elif SUCCESS == 1:
                self.logwindow.config(text='Die Datei ist noch geöffnet. Bitte schließen Sie die Datei', foreground='red')
            else:
                self.logwindow.config(text='Excel konnte nicht erstellt werden', foreground='red')
        

    def runExcel(self):
        file = self.file_path[0]
        self.xlsx_file_name = file.replace('.txt', '.xlsx')

        # Klassenliste laden
        with open(file,'r') as file:
            namen = file.readlines()

        # Variable als Container fuer alle Schüler
        try:
            self.Klasse = []
            for name in namen:
                m = re_search(r'(\w+\s*\w*\.*)\s*\t+(\w+\s*\w*)', name)
                if m.group(2) == '' or m.group(1) == '':
                    self.logwindow(text='Der/Die SchülerIn ' + m.group(2) + ' ' + m.group(1) + ' konnte nicht richtig eingelesen werden.')
                else:
                    self.Klasse.append(Schueler(m.group(2), m.group(1)))
        except Exception as e:
            self.logwindow.config(text='Fehler beim Auslesen der Klassenliste')


        # Excel-File erstellen
        self.wb = Workbook()
        ws = self.wb.active
        if self.benennung.get() != '':
        	self.worksheet_title = self.benennung.get()
        else:
        	self.worksheet_title = 'Klassenübersicht'
        ws.title = self.worksheet_title
        ws.sheet_properties.tabColor = "1072BA"

        # Querformat
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

        lastname_title = ws[columns[Spalte_erste_Aufgabe-5]+str(Zeile_erster_Schueler-1)]
        lastname_title.value = self.LASTNAME
        lastname_title.font = Font(bold=True)
        firstname_title = ws[columns[Spalte_erste_Aufgabe-4]+str(Zeile_erster_Schueler-1)]
        firstname_title.value = self.FIRSTNAME
        firstname_title.font = Font(bold=True)
        # Rahmen erstellen
        makeBorder(ws,lastname_title.coordinate + ':' + firstname_title.coordinate)

        # ---------------------------------------------------------------------------------------------------------------
        # Namen aller Schueler einstragen
        start_row = Zeile_erster_Schueler
        nachname_column = Spalte_erste_Aufgabe-5
        vorname_column = Spalte_erste_Aufgabe-4
        for s in range(len(self.Klasse)):
            # Nachname
            ws[columns[nachname_column] + str(start_row+s)] = self.Klasse[s].nachname
            # Vorname
            ws[columns[vorname_column] + str(start_row+s)] = self.Klasse[s].vorname

        # Die Spalten mit Rahmen versehen
        makeBorder(ws,columns[nachname_column]+str(start_row)+':'+columns[vorname_column]+str(start_row+len(self.Klasse)))

        # ---------------------------------------------------------------------------------------------------------------
        # Spalten fuer die Noten
        akt_cell = ws[columns[Spalte_erste_Aufgabe-3] + str(Zeile_erster_Schueler-1)]
        akt_cell.value = 'Note'
        akt_cell.font = Font(bold=True)
        akt_cell.alignment = Alignment(horizontal="center")
        makeBorder(ws,akt_cell.coordinate)
        ws.merge_cells(columns[Spalte_erste_Aufgabe-3] + str(Zeile_erster_Schueler-1)+':'+columns[Spalte_erste_Aufgabe-1] + str(Zeile_erster_Schueler-1))

        # ---------------------------------------------------------------------------------------------------------------
        # Aufgaben anlegen
        start_row = Zeile_erster_Schueler-3
        start_column = Spalte_erste_Aufgabe

        tasks = self.tasks
        needed_cols = sum([tasks[i].getNumCols() for i in range(len(tasks))])
        num_tasks = len(tasks)
        punkte_cols = []

        for i in range(num_tasks):
            akt_row = start_row
            akt_task = ws[columns[start_column] + str(start_row)]
            akt_task.value = tasks[i].getTitel()
            akt_task.font = Font(bold=True)
            makeBorder(ws,akt_task.coordinate)
            akt_task.alignment = Alignment(horizontal="center", vertical="center")
            ws.merge_cells(columns[start_column] + str(start_row) + ':' + columns[start_column + tasks[i].getNumCols()-1] + str(start_row))

            if tasks[i].getType() == 'Aufgabe':
                akt_cell = ws[columns[start_column] + str(akt_row+1)]
                akt_cell.value = 'BE'
                akt_cell.alignment = Alignment(horizontal="center", vertical="center")
                akt_cell.border = Border(left=thin_black)
                akt_cell = ws[columns[start_column] + str(akt_row+2)]
                akt_cell.value = tasks[i].getBE()
                akt_cell.alignment = Alignment(horizontal="center", vertical="center")
                akt_cell.border = Border(left=thin_black, bottom=thin_black)

                akt_cell = ws[columns[start_column+1] + str(akt_row + 1)]
                akt_cell.value = 'Gewichtung'
                akt_cell.alignment = Alignment(horizontal="center", vertical="center")
                akt_cell.border = Border(right=thin_black)
                akt_cell = ws[columns[start_column+i*2+1] + str(akt_row + 2)]
                akt_cell.value = tasks[i].getGewichtung()
                akt_cell.alignment = Alignment(horizontal="center", vertical="center")
                akt_cell.border = Border(right=thin_black, bottom=thin_black)

                # Berechnung in die Spalte mit Gewichtung
                for t in range(len(self.Klasse)):
                    akt_cell = ws[columns[start_column+1] + str(Zeile_erster_Schueler+t)]
                    akt_cell.value = '=' + columns[start_column] + str(Zeile_erster_Schueler+t) + '*$' + columns[start_column+1] + '$' + str(akt_row + 2)
                    akt_cell.alignment = Alignment(horizontal="center")

                    akt_cell = ws[columns[start_column] + str(Zeile_erster_Schueler+t)]
                    akt_cell.value = ''
                    akt_cell.alignment = Alignment(horizontal='center')        
                makeBorder(ws,columns[start_column] + str(Zeile_erster_Schueler)+':'+columns[start_column+1] + str(Zeile_erster_Schueler+len(self.Klasse)))

            elif tasks[i].getType() == 'Textproduktion':
                akt_cell = ws[columns[start_column] + str(akt_row+1)]
                akt_cell.value = 'Inhalt'
                akt_cell.alignment = Alignment(horizontal="center", vertical="center")
                akt_cell.border = Border(left=thin_black)
                akt_cell = ws[columns[start_column] + str(akt_row+2)]
                akt_cell.value = tasks[i].getInhalt()
                akt_cell.alignment = Alignment(horizontal="center", vertical="center")
                akt_cell.border = Border(left=thin_black, bottom=thin_black)

                akt_cell = ws[columns[start_column+1] + str(akt_row+1)]
                akt_cell.value = 'Sprache'
                akt_cell.alignment = Alignment(horizontal="center", vertical="center")
                akt_cell.border = Border(left=thin_black)
                akt_cell = ws[columns[start_column+1] + str(akt_row+2)]
                akt_cell.value = tasks[i].getSprache()
                akt_cell.alignment = Alignment(horizontal="center", vertical="center")
                akt_cell.border = Border(left=thin_black, bottom=thin_black)

                akt_cell = ws[columns[start_column+2] + str(akt_row+1)]
                akt_cell.value = 'Gewichtung'
                akt_cell.alignment = Alignment(horizontal="center", vertical="center")
                akt_cell.border = Border(left=thin_black)
                akt_cell = ws[columns[start_column+2] + str(akt_row+2)]
                akt_cell.value = tasks[i].getGewichtung()
                akt_cell.alignment = Alignment(horizontal="center", vertical="center")
                akt_cell.border = Border(left=thin_black, bottom=thin_black)

                # Berechnung in die Spalte mit Gewichtung
                for t in range(len(self.Klasse)):
                    akt_cell = ws[columns[start_column+2] + str(Zeile_erster_Schueler+t)]
                    akt_cell.value = '=(' + columns[start_column] + str(Zeile_erster_Schueler+t) + '+' + columns[start_column+1] + str(Zeile_erster_Schueler+t) + ')*$' + columns[start_column+2] + '$' + str(akt_row + 2)
                    akt_cell.alignment = Alignment(horizontal="center")

                    akt_cell = ws[columns[start_column] + str(Zeile_erster_Schueler+t)]
                    akt_cell.value = ''
                    akt_cell.alignment = Alignment(horizontal='center')  

                    akt_cell = ws[columns[start_column+1] + str(Zeile_erster_Schueler+t)]
                    akt_cell.value = ''
                    akt_cell.alignment = Alignment(horizontal='center')       
                makeBorder(ws,columns[start_column] + str(Zeile_erster_Schueler)+':'+columns[start_column+2] + str(Zeile_erster_Schueler+len(self.Klasse)))

            # Breite der Spalten anpassen
            for s in range(tasks[i].getNumCols()-1):
                ws.column_dimensions[columns[start_column+s]].width = 10
            ws.column_dimensions[columns[start_column+tasks[i].getNumCols()-1]].width = 15

            # Neue Startspalte
            punkte_cols.append(columns[start_column+tasks[i].getNumCols()-1])
            start_column = start_column + tasks[i].getNumCols()

        # Gesamt
        start_row = Zeile_erster_Schueler-3
        spalte_gesamt = Spalte_erste_Aufgabe + needed_cols 
        start_column = spalte_gesamt

        akt_cell = ws[columns[spalte_gesamt] + str(start_row)]
        akt_cell.value = 'Gesamt'
        akt_cell.font = Font(bold=True)
        makeBorder(ws,akt_cell.coordinate)
        akt_cell.alignment = Alignment(horizontal="center", vertical="bottom")
        ws.merge_cells(columns[spalte_gesamt] + str(start_row) + ':' + columns[spalte_gesamt] + str(start_row+1))

        # Berechnung der Gesamtpunktzahl fuer jeden Schueler
        start_row = Zeile_erster_Schueler
        for i in range(len(self.Klasse)):
            calc_string = '='
            for t in range(len(tasks)):
                calc_string = calc_string + punkte_cols[t] + str(start_row+i) + '+'

            calc_string = calc_string[0:len(calc_string)-1] # letztes + entfernen

            ws[columns[spalte_gesamt] + str(start_row+i)] = calc_string

        makeBorder(ws,columns[Spalte_erste_Aufgabe+needed_cols]+str(start_row)+':'+columns[Spalte_erste_Aufgabe+needed_cols]+str(start_row+len(self.Klasse)-1))

        # Gesamtpunktzahl
        calc_string = '='
        for t in range(len(tasks)):
            start_column = columns.find(punkte_cols[t])
            if tasks[t].getNumCols() == 2:
                calc_string = calc_string + columns[start_column-1] + str(Zeile_erster_Schueler-1) + '*' + columns[start_column]  + str(Zeile_erster_Schueler-1) + '+'
            elif tasks[t].getNumCols() == 3:
                calc_string = calc_string + '(' + columns[start_column-2] + str(Zeile_erster_Schueler-1) + '+' + columns[start_column-1] + str(Zeile_erster_Schueler-1) + ')*' + columns[start_column]  + str(Zeile_erster_Schueler-1) + '+'
        calc_string = calc_string[0:len(calc_string)-1]

        cell_gesamtpunktzahl = columns[Spalte_erste_Aufgabe+needed_cols] + str(Zeile_erster_Schueler-1)
        akt_cell = ws[cell_gesamtpunktzahl]
        akt_cell.value = calc_string
        akt_cell.font = Font(bold=True)
        makeBorder(ws,akt_cell.coordinate)

        # Übersicht
        spalte_uebersicht_noten = spalte_gesamt + 2
        stufung = [100, 87.5, 75, 62.5, 50, 33, 0]

        # Ueberschriften
        akt_cell = ws[columns[spalte_uebersicht_noten]+str(Zeile_erster_Schueler-1)]
        akt_cell.value = 'Note'
        akt_cell.font = Font(bold=True)
        akt_cell = ws[columns[spalte_uebersicht_noten+1]+str(Zeile_erster_Schueler-1)]
        akt_cell.value = 'Prozentsatz'
        akt_cell.font = Font(bold=True)
        akt_cell = ws[columns[spalte_uebersicht_noten+2]+str(Zeile_erster_Schueler-1)]
        akt_cell.value = 'Punkte'
        akt_cell.font = Font(bold=True)
        akt_cell.alignment = Alignment(horizontal="center")
        ws.merge_cells(columns[spalte_uebersicht_noten+2]+str(Zeile_erster_Schueler-1)+':'+columns[spalte_uebersicht_noten+4]+str(Zeile_erster_Schueler-1))
        akt_cell = ws[columns[spalte_uebersicht_noten+5]+str(Zeile_erster_Schueler-1)]
        akt_cell.value = 'Anzahl'
        akt_cell.font = Font(bold=True)
        makeBorder(ws,columns[spalte_uebersicht_noten]+str(Zeile_erster_Schueler-1)+':'+columns[spalte_uebersicht_noten+5]+str(Zeile_erster_Schueler-1))

        str_calc_noten = '=IF(NOT(ISNUMBER({0})),"",'
        str_calc_plus = '=IF(NOT(ISNUMBER({0})),"",IF(OR('
        str_calc_minus = str_calc_plus
        cells_mit_noten, cells_mit_anzahl = [], []
        for note in range(1, 6+1):
            akt_row = Zeile_erster_Schueler+note-1
            akt_cell = ws[columns[spalte_uebersicht_noten]+str(akt_row)]
            akt_cell.value = note
            akt_cell.border = Border(left=thin_black)
            akt_cell.alignment = Alignment(horizontal="center")

            # Prozentsatz
            akt_cell = ws[columns[spalte_uebersicht_noten+1]+str(akt_row)]
            akt_cell.value = stufung[note]
            akt_cell.alignment = Alignment(horizontal="center")

            # Punktzahlbereich
            akt_cell = ws[columns[spalte_uebersicht_noten+2]+str(akt_row)]
            if note == 1:
                akt_cell.value = '=' + cell_gesamtpunktzahl
            else:
                akt_cell.value = '=ROUNDDOWN(' + cell_gesamtpunktzahl + '*' + columns[spalte_uebersicht_noten+1]+str(akt_row-1) + '/100*2,0)/2-0.5'
            akt_cell.alignment = Alignment(horizontal="right")
            akt_cell = ws[columns[spalte_uebersicht_noten + 3] + str(akt_row)]
            akt_cell.value = '-'
            akt_cell.alignment = Alignment(horizontal="center")
            akt_cell = ws[columns[spalte_uebersicht_noten + 4] + str(akt_row)]
            akt_cell.value = '=ROUNDDOWN(' + cell_gesamtpunktzahl + '*' + columns[spalte_uebersicht_noten+1]+str(akt_row) + '/100*2,0)/2'
            akt_cell.alignment = Alignment(horizontal="left")

            # Anzahl der Schueler mit den jeweiligen Noten
            akt_cell = ws[columns[spalte_uebersicht_noten + 5] + str(akt_row)]
            akt_cell.value = '=COUNTIF(' + columns[Spalte_erste_Aufgabe-2] + str(Zeile_erster_Schueler) + ':' + columns[Spalte_erste_Aufgabe-2] + str(Zeile_erster_Schueler+len(self.Klasse)-1) + ',' + columns[spalte_uebersicht_noten] + str(akt_row) + ')'

            # String fuer die Berechnung der Noten erstellen
            str_calc_noten = str_calc_noten + 'IF({0}>=$' + columns[spalte_uebersicht_noten + 4] + '$' + str(akt_row) + ',' + str(note) + ','
            str_calc_plus = str_calc_plus + '{0}=$' + columns[spalte_uebersicht_noten + 2] + '$' + str(akt_row) + ','
            str_calc_minus = str_calc_minus + '{0}=$' + columns[spalte_uebersicht_noten + 4] + '$' + str(akt_row) + ','

            cells_mit_noten.append([columns[spalte_uebersicht_noten],str(akt_row)])
            cells_mit_anzahl.append([columns[spalte_uebersicht_noten + 5] , str(akt_row)])

        for i in range(6):
            ws.column_dimensions[columns[spalte_uebersicht_noten+i]].width = 8
        makeBorder(ws,columns[spalte_uebersicht_noten]+str(Zeile_erster_Schueler)+':'+columns[spalte_uebersicht_noten+5]+str(Zeile_erster_Schueler+6))


        str_calc_noten = str_calc_noten + ')))))))'
        str_calc_plus = str_calc_plus[0:len(str_calc_plus)-1] + '),"+",""))'
        str_calc_minus = str_calc_minus[0:len(str_calc_minus)-1] + '),"-",""))'

        # Berechnung der Noten anhand der Stufungen
        for s in range(len(self.Klasse)):
            akt_cell = ws[columns[Spalte_erste_Aufgabe-2]+str(Zeile_erster_Schueler+s)]
            schueler_gesamt = cell_gesamtpunktzahl[0] + str(Zeile_erster_Schueler+s)

            # Note
            akt_cell.value = str_calc_noten.format(schueler_gesamt)
            akt_cell.font = Font(bold=True)
            akt_cell.alignment = Alignment(horizontal='center')

            # Plus
            akt_cell = ws[columns[Spalte_erste_Aufgabe-3]+str(Zeile_erster_Schueler+s)]
            akt_cell.value = str_calc_plus.format(schueler_gesamt)
            akt_cell.font = Font(bold=True)
            akt_cell.alignment = Alignment(horizontal='center')

            # Minus
            akt_cell = ws[columns[Spalte_erste_Aufgabe - 1] + str(Zeile_erster_Schueler + s)]
            akt_cell.value = str_calc_minus.format(schueler_gesamt)
            akt_cell.font = Font(bold=True)
            akt_cell.alignment = Alignment(horizontal='center')

        makeBorder(ws,columns[Spalte_erste_Aufgabe-3]+str(Zeile_erster_Schueler)+':'+columns[Spalte_erste_Aufgabe-1]+str(Zeile_erster_Schueler+len(self.Klasse)))

        # conditional formatting
        cells_noten = columns[Spalte_erste_Aufgabe-2]+str(Zeile_erster_Schueler)+':'+columns[Spalte_erste_Aufgabe-2]+str(Zeile_erster_Schueler+len(self.Klasse)-1)
        #ws.conditional_formatting.add2ColorScale(cells_noten, 'min', 1, 'FFAA0000', 'max', 6, 'FF00AA00')
        rule = {cells_noten: [{'type': 'colorScale', 'priority': 13,'colorScale': {'cfvo': [{'type': 'min'}, {'type': 'max'}], 'color': [Color('FFFF7128'), Color('FFFFEF9C')]}}]}

        #ws.conditional_formatting.add(rule)
        ws.conditional_formatting.add(cells_noten,ColorScaleRule(start_type='num', start_value=1, start_color=Color('00b034'),mid_type='percentile', mid_value=50, mid_color=Color('ffc000'), end_type='num', end_value=6, end_color=Color('ff0000')))


        # Spaltenbreiste anpassen
        ws.column_dimensions[columns[Spalte_erste_Aufgabe-1]].width = 3
        ws.column_dimensions[columns[Spalte_erste_Aufgabe-2]].width = 5
        ws.column_dimensions[columns[Spalte_erste_Aufgabe-3]].width = 3
        ws.column_dimensions[columns[Spalte_erste_Aufgabe-4]].width = 25
        ws.column_dimensions[columns[Spalte_erste_Aufgabe-5]].width = 25


        # Diagramm
        self.chart1 = BarChart()
        self.chart1.type = "col"
        self.chart1.style = 10
        self.chart1.title = None
        self.chart1.y_axis.title = 'Anzahl'
        self.chart1.x_axis.title = 'Note'
        self.chart1.legend = None

        data = Reference(ws, min_col=int(columns.find(cells_mit_anzahl[0][0]))+1, min_row=int(cells_mit_anzahl[0][-1]), max_row=int(cells_mit_anzahl[-1][-1]), max_col=int(columns.find(cells_mit_anzahl[-1][0]))+1)
        ref = Reference(ws, min_col=int(columns.find(cells_mit_noten[0][0]))+1, min_row=int(cells_mit_noten[0][-1]), max_row=int(cells_mit_noten[-1][-1]), max_col=int(columns.find(cells_mit_noten[-1][0]))+1)
        self.chart1.add_data(data, titles_from_data=False)
        self.chart1.set_categories(ref)
        self.chart1.shape = 4
        ws.add_chart(self.chart1, cells_mit_noten[0][0]+str(int(cells_mit_noten[-1][-1])+2))

        # Speichern
        try:
            self.wb.save(self.xlsx_file_name)
            self.showButtonOpen(self.xlsx_file_name)
            return 0
        except PermissionError:
            print('Die Datei ist noch geöffnet und muss erst geschlossen werden!')
            return 1
        except Exception as e:
            print('Etwas ist fehlgeschlagen')
            print(e)
            return 100

    def showButtonOpen(self,file_to_open):
        os_system('start EXCEL.EXE "' + file_to_open + '"')



class Schueler:
    def __init__(self, nachname, vorname):
        self.nachname = nachname
        self.vorname = vorname

    def getNachname(self):
        return self.nachname

    def getVorname(self):
        return self.vorname

    def getName(self):
        return self.vorname + ' ' + self.nachname

class Aufgabe:
    def __init__(self, BE, gewichtung=1, titel='Aufgabe'):
        self.BE = BE
        self.gewichtung = gewichtung
        self.titel = titel
        self.type = 'Aufgabe'

    def getBE(self):
        return self.BE

    def getGewichtung(self):
        return self.gewichtung

    def getTitel(self):
        return self.titel

    def getType(self):
        return self.type

    def setBE(self,BE):
        self.BE = BE

    def setGewichtung(self, gewichtung):
        self.gewichtung = gewichtung

    def setTitel(self, titel):
        self.titel = str(titel)

    def getNumCols(self):
        return 2

class Textproduktion(Aufgabe):
    def __init__(self,inhalt,sprache,gewichtung, titel='Textproduktion'):
        self.inhalt = inhalt
        self.sprache = sprache
        self.gewichtung = gewichtung
        self.titel = titel
        self.type = 'Textproduktion'

    def getInhalt(self):
        return self.inhalt

    def getSprache(self):
        return self.sprache

    def getBE(self):
        return self.inhalt + self.sprache

    def setTitel(self, titel):
        self.titel = str(titel)

    def getNumCols(self):
        return 3



def makeBorder(worksheet,cells):

    # if no ':' was found
    if cells.find(':') == -1:
        m = re_search(r'([A-Z]+\d+)',cells)
        if m.group(1)!='':
            worksheet[cells].border = Border(top=thin_black, left=thin_black, right=thin_black, bottom=thin_black)
        else:
            return
    else:
        m = re_search(r'([A-Z]+\d+):([A-Z]+\d+)',cells)  
        start_cell = m.group(1)
        end_cell = m.group(2)

        start_letter = re_search(r'[A-Z]+',start_cell).group(0)
        end_letter = re_search(r'[A-Z]+',end_cell).group(0)
        start_number = re_search(r'[0-9]+',start_cell).group(0)
        end_number = re_search(r'[0-9]+',end_cell).group(0)

        start_column = columns.find(start_letter)
        end_column = columns.find(end_letter)
        if start_column > end_column:
            start_column, end_column = end_column, start_column
        num_columns = end_column-start_column

        start_row = int(start_number)
        end_row = int(end_number)
        if start_row > end_row:
            start_row, end_row = end_row, start_row
        num_rows = end_row - start_row

        if num_columns < 0 or num_rows < 0:
            return

        # Wenn alle Zellen in einer Zeile sind
        if num_rows == 0:
            for c in range(start_column,end_column+1,1):
                if c == start_column:
                    worksheet[columns[c]+str(start_row)].border = Border(top=thin_black, left=thin_black, bottom=thin_black)
                elif c == end_column:
                    worksheet[columns[c]+str(start_row)].border = Border(top=thin_black, bottom=thin_black, right=thin_black)
                else:
                    worksheet[columns[c]+str(start_row)].border = Border(top=thin_black, bottom=thin_black)
        # Wenn alle Zellen in einer Spalte sind
        elif num_columns == 0:
            for r in range(start_row,end_row+1,1):
                if r == start_row:
                    worksheet[columns[start_column]+str(r)].border = Border(top=thin_black, left=thin_black, right=thin_black)
                elif r == end_row:
                    worksheet[columns[start_column]+str(r)].border = Border(left=thin_black, bottom=thin_black, right=thin_black)
                else:
                    worksheet[columns[start_column]+str(r)].border = Border(left=thin_black, right=thin_black)
        # Wenn die Auswahl ueber mehrere Zeilen und Spalten geht
        else:
            for r in range(num_rows):
                for c in range(num_columns+1):
                    akt_cell = columns[start_column+c]+str(start_row+r)

                    # erste Zeile
                    if r == 0:
                        # erste Spalte
                        if c == 0:
                            worksheet[akt_cell].border = Border(left=thin_black, top=thin_black)
                        # letzte Spalte
                        elif c == num_columns:
                            worksheet[akt_cell].border = Border(right=thin_black, top=thin_black)
                        # zwischen erster und letzter Spalte
                        else:
                            worksheet[akt_cell].border = Border(top=thin_black)
                    # letzte Zeile
                    elif r == num_rows-1:
                        # erste Spalte
                        if c == 0:
                            worksheet[akt_cell].border = Border(left=thin_black, bottom=thin_black)
                        # letzte Spalte
                        elif c == num_columns:
                            worksheet[akt_cell].border = Border(right=thin_black, bottom=thin_black)
                        # zwischen erster und letzter Spalte
                        else:
                            worksheet[akt_cell].border = Border(bottom=thin_black)
                    # zwischen erster und letzter Zeile
                    else:
                        # erste Spalte
                        if c == 0:
                            worksheet[akt_cell].border = Border(left=thin_black)
                        # letzte Spalte
                        elif c == num_columns:
                            worksheet[akt_cell].border = Border(right=thin_black)


# Zeilen und Spalten definieren
Zeile_erster_Schueler = 5
Spalte_erste_Aufgabe = 6 # siebte Spalte

# Cell border styles
thin_black = Side(border_style="thin", color="000000")


columns = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'

all_entries = []
